#!/usr/bin/env python3
"""
Extract test data from the roadmap PDF and create Excel file
Based on the PS 2.0 One Year Roadmap (Feb-Aug 2026)
"""

import json
from datetime import datetime, timedelta

# Define test data based on the roadmap
test_tasks = [
    # Hotel Supply: Connectivity and Local Contracts (Jan-Jul)
    {
        "id": "1",
        "name": "Hotel Supply: Connectivity and Local Contracts",
        "startDate": "2026-01-15",
        "endDate": "2026-07-31",
        "dependencies": []
    },
    # Pricing rules (Jan-Aug)
    {
        "id": "2",
        "name": "Pricing rules - BAU",
        "startDate": "2026-01-01",
        "endDate": "2026-08-31",
        "dependencies": []
    },
    # Local Contracts Go Live (May)
    {
        "id": "3",
        "name": "Local Contracts Go Live (OJ + Darwin Flow)",
        "startDate": "2026-05-01",
        "endDate": "2026-05-31",
        "dependencies": ["1"]
    },
    # Phase 2 Destinations Enabling (Brussels, Copenhagen, Ibiza) - Feb-Mar
    {
        "id": "4",
        "name": "Phase 2 Destinations Enabling - Brussels, Copenhagen, Ibiza",
        "startDate": "2026-02-01",
        "endDate": "2026-03-15",
        "dependencies": []
    },
    # Phase 3 Destinations Enabling (Mar 23rd) - Amsterdam, Athens, Dublin, Glasgow, Manchester
    {
        "id": "5",
        "name": "Phase 3 Destinations Enabling - Amsterdam, Athens, Dublin, Glasgow, Manchester",
        "startDate": "2026-03-01",
        "endDate": "2026-03-23",
        "dependencies": ["4"]
    },
    # Phase 4 Destinations Enabling (Apr 23rd) - Barcelona, Istanbul, Malta
    {
        "id": "6",
        "name": "Phase 4 Destinations Enabling - Barcelona, Istanbul, Malta",
        "startDate": "2026-04-01",
        "endDate": "2026-04-23",
        "dependencies": ["5"]
    },
    # Phase 5 Destinations Enabling (May 26th) - Boston, Chicago, Jersey, Marrakech, Miami
    {
        "id": "7",
        "name": "Phase 5 Destinations Enabling - Boston, Chicago, Jersey, Marrakech, Miami",
        "startDate": "2026-05-01",
        "endDate": "2026-05-26",
        "dependencies": ["6"]
    },
    # Tech / Platform work
    {
        "id": "8",
        "name": "Tech / Platform work",
        "startDate": "2026-03-15",
        "endDate": "2026-08-15",
        "dependencies": []
    },
    # Automate product setup
    {
        "id": "9",
        "name": "Automate product setup",
        "startDate": "2026-04-01",
        "endDate": "2026-07-31",
        "dependencies": ["8"]
    },
    # PS: Codegen Gaps
    {
        "id": "10",
        "name": "PS: Codegen Gaps",
        "startDate": "2026-05-01",
        "endDate": "2026-08-31",
        "dependencies": []
    },
    # Promotions calendar
    {
        "id": "11",
        "name": "Promotions calendar",
        "startDate": "2026-04-15",
        "endDate": "2026-06-30",
        "dependencies": ["8"]
    },
    # Remove Dynamic Profile Creation completely
    {
        "id": "12",
        "name": "Remove Dynamic Profile Creation completely or remove from Mik's team",
        "startDate": "2026-03-01",
        "endDate": "2026-04-15",
        "dependencies": []
    },
    # Automate Agresso set up for supplier info
    {
        "id": "13",
        "name": "Automate Agresso set up for supplier info (to migrate work from Mik's team to WNS)",
        "startDate": "2026-04-01",
        "endDate": "2026-06-15",
        "dependencies": ["9"]
    },
    # Product Setup for Transfers
    {
        "id": "14",
        "name": "Product Setup for Transfers",
        "startDate": "2026-05-15",
        "endDate": "2026-07-15",
        "dependencies": ["9"]
    },
    # Set up & content for Cars
    {
        "id": "15",
        "name": "Set up & content for Cars",
        "startDate": "2026-05-01",
        "endDate": "2026-06-30",
        "dependencies": ["9"]
    },
    # Detailed plan for locally held contract dual run go live (in OJ)
    {
        "id": "16",
        "name": "Detailed plan for locally held contract dual run go live (in OJ)",
        "startDate": "2026-03-15",
        "endDate": "2026-04-30",
        "dependencies": ["1"]
    },
    # Revise usefulness of Splunk dashboards
    {
        "id": "17",
        "name": "Revise usefulness of Splunk dashboards",
        "startDate": "2026-04-01",
        "endDate": "2026-05-15",
        "dependencies": ["8"]
    },
    # Set up monitoring and alerting
    {
        "id": "18",
        "name": "Set up monitoring and alerting",
        "startDate": "2026-05-01",
        "endDate": "2026-06-30",
        "dependencies": ["17"]
    },
    # Ensure observability is sufficient
    {
        "id": "19",
        "name": "Ensure observability is sufficient to allow troubleshooting of alerts",
        "startDate": "2026-06-01",
        "endDate": "2026-07-31",
        "dependencies": ["18"]
    },
    # Set up and progressively extend automated regression testing in TBX
    {
        "id": "20",
        "name": "Set up and progressively extend automated regression testing in TBX",
        "startDate": "2026-04-15",
        "endDate": "2026-08-31",
        "dependencies": ["8"]
    },
    # Add Vegas as a Destination
    {
        "id": "21",
        "name": "Add Vegas as a Destination",
        "startDate": "2026-06-15",
        "endDate": "2026-07-31",
        "dependencies": ["7"]
    },
    # LC - TravelGate Integration (OJ/TBX)
    {
        "id": "22",
        "name": "LC - TravelGate Integration (OJ/TBX)",
        "startDate": "2026-05-01",
        "endDate": "2026-08-31",
        "dependencies": ["8"]
    },
    # Training for Destinations team
    {
        "id": "23",
        "name": "Training for Destinations team",
        "startDate": "2026-07-01",
        "endDate": "2026-08-15",
        "dependencies": ["7"]
    },
    # Training for WNS
    {
        "id": "24",
        "name": "Training for WNS",
        "startDate": "2026-07-15",
        "endDate": "2026-08-31",
        "dependencies": ["13"]
    },
    # Training for Pricing
    {
        "id": "25",
        "name": "Training for Pricing",
        "startDate": "2026-07-15",
        "endDate": "2026-08-31",
        "dependencies": ["2"]
    },
    # Training for BAU testers
    {
        "id": "26",
        "name": "Training for BAU testers",
        "startDate": "2026-07-01",
        "endDate": "2026-08-15",
        "dependencies": ["20"]
    },
    # Training for Operations (Local)
    {
        "id": "27",
        "name": "Training for Operations (Local)",
        "startDate": "2026-07-15",
        "endDate": "2026-08-31",
        "dependencies": ["3"]
    },
    # Training for Contentful
    {
        "id": "28",
        "name": "Training for Contentful (Local/PS)",
        "startDate": "2026-07-01",
        "endDate": "2026-08-15",
        "dependencies": ["9"]
    },
    # Headline Price + Local Fees API feeds
    {
        "id": "29",
        "name": "Headline Price + Local Fees API feeds for Hotel Only and Packages (DP Flight + Hotel)",
        "startDate": "2026-02-01",
        "endDate": "2026-04-30",
        "dependencies": ["2"]
    },
    # Modular Price Comparison: Dilution Rules
    {
        "id": "30",
        "name": "Modular Price Comparison: Dilution Rules - Ability to flex package price",
        "startDate": "2026-03-01",
        "endDate": "2026-05-31",
        "dependencies": ["29"]
    },
    # Markup Scheme / Discount Scheme Adjustment
    {
        "id": "31",
        "name": "Markup Scheme / Discount Scheme Adjustment based on bundled or individual component cost",
        "startDate": "2026-04-01",
        "endDate": "2026-06-30",
        "dependencies": ["30"]
    },
    # Modular Price Comparison: Close the Gap
    {
        "id": "32",
        "name": "Modular Price Comparison: Close the Gap",
        "startDate": "2026-03-15",
        "endDate": "2026-05-15",
        "dependencies": ["29"]
    },
    # Hotel Content Model
    {
        "id": "33",
        "name": "Hotel Content Model",
        "startDate": "2026-04-01",
        "endDate": "2026-07-31",
        "dependencies": ["9"]
    },
    # Product Setup process optimisation
    {
        "id": "34",
        "name": "Product Setup process optimisation",
        "startDate": "2026-05-01",
        "endDate": "2026-07-15",
        "dependencies": ["9"]
    },
    # CMA Headline Price
    {
        "id": "35",
        "name": "CMA Headline Price",
        "startDate": "2026-05-15",
        "endDate": "2026-08-31",
        "dependencies": ["29"]
    },
    # Flight Dilution rules
    {
        "id": "36",
        "name": "Flight Dilution rules",
        "startDate": "2026-04-15",
        "endDate": "2026-07-31",
        "dependencies": ["30"]
    },
    # Bundles core pricing rules
    {
        "id": "37",
        "name": "Bundles core pricing rules + pricing and markup enhancements (integration of private fares)",
        "startDate": "2026-04-01",
        "endDate": "2026-08-31",
        "dependencies": ["31"]
    },
    # Minimum Margin Protection
    {
        "id": "38",
        "name": "Minimum Margin Protection",
        "startDate": "2026-05-15",
        "endDate": "2026-08-31",
        "dependencies": ["37"]
    },
    # Automated content loading solution
    {
        "id": "39",
        "name": "Automated content loading solution - Automation of Product Loading using CodeGen AI",
        "startDate": "2026-05-01",
        "endDate": "2026-08-15",
        "dependencies": ["10"]
    },
    # IBS Connectivity & Add Vegas
    {
        "id": "40",
        "name": "IBS Connectivity & Add Vegas as a destination",
        "startDate": "2026-06-01",
        "endDate": "2026-08-31",
        "dependencies": ["21"]
    },
    # LC - Smart Contract Loading
    {
        "id": "41",
        "name": "LC - Smart Contract Loading",
        "startDate": "2026-06-01",
        "endDate": "2026-08-31",
        "dependencies": ["33"]
    },
    # Get relevant hotel recommendations - Discovery
    {
        "id": "42",
        "name": "Get relevant hotel recommendations - Discovery Phase",
        "startDate": "2026-04-01",
        "endDate": "2026-05-31",
        "dependencies": []
    },
    # Get relevant hotel recommendations - Implementation
    {
        "id": "43",
        "name": "Get relevant hotel recommendations - Implementation",
        "startDate": "2026-06-01",
        "endDate": "2026-08-31",
        "dependencies": ["42"]
    },
    # Surface Best Room Offer
    {
        "id": "44",
        "name": "Surface Best Room Offer - Prepare Discovery",
        "startDate": "2026-04-15",
        "endDate": "2026-06-15",
        "dependencies": []
    },
    # LC - Supplier Contracts & Communication (PDF)
    {
        "id": "45",
        "name": "LC - Supplier Contracts & Communication (PDF)",
        "startDate": "2026-03-01",
        "endDate": "2026-07-31",
        "dependencies": ["1"]
    },
    # LC - Supplier Contracts & Communication (Template)
    {
        "id": "46",
        "name": "LC - Supplier Contracts & Communication (Template - Blocker)",
        "startDate": "2026-04-01",
        "endDate": "2026-06-30",
        "dependencies": ["45"]
    },
    # Data Migration & Inventory
    {
        "id": "47",
        "name": "Data Migration & Inventory",
        "startDate": "2026-05-01",
        "endDate": "2026-08-31",
        "dependencies": ["46"]
    },
    # Create Tritium IDs in Travelbox and Hotel loading
    {
        "id": "48",
        "name": "Create Tritium IDs in Travelbox and Hotel loading (10,500 hotels live by end July)",
        "startDate": "2026-05-15",
        "endDate": "2026-07-31",
        "dependencies": ["47"]
    },
    # LC - Finance & Agresso
    {
        "id": "49",
        "name": "LC - Finance & Agresso",
        "startDate": "2026-05-01",
        "endDate": "2026-08-31",
        "dependencies": ["13"]
    },
    # IBS Connectivity - Mik
    {
        "id": "50",
        "name": "IBS Connectivity - Setup",
        "startDate": "2026-04-01",
        "endDate": "2026-07-31",
        "dependencies": ["8"]
    },
    # Hotel Scaling - Manual Phase
    {
        "id": "51",
        "name": "Manual Hotel Scaling",
        "startDate": "2026-02-01",
        "endDate": "2026-04-30",
        "dependencies": []
    },
    # Hotel Scaling - Automated Phase
    {
        "id": "52",
        "name": "Automated Hotel Scaling",
        "startDate": "2026-05-01",
        "endDate": "2026-08-31",
        "dependencies": ["51", "39"]
    },
]

def save_as_json(filename='tasks.json'):
    """Save test data as JSON"""
    with open(filename, 'w') as f:
        json.dump(test_tasks, f, indent=2)
    print(f"✓ Saved {len(test_tasks)} tasks to {filename}")

def save_as_excel(filename='roadmap_test_data.xlsx'):
    """Save test data as Excel file"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed. Installing...")
        import subprocess
        subprocess.check_call(['pip3', 'install', 'openpyxl'])
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Roadmap Tasks"
    
    # Define styles
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Write headers
    headers = ["Task ID", "Task Name", "Start Date", "End Date", "Duration (Days)", "Dependencies"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Write data
    for row, task in enumerate(test_tasks, 2):
        start = datetime.strptime(task['startDate'], '%Y-%m-%d')
        end = datetime.strptime(task['endDate'], '%Y-%m-%d')
        duration = (end - start).days
        dependencies_str = ", ".join(task['dependencies']) if task['dependencies'] else "-"
        
        cells = [
            ws.cell(row=row, column=1, value=task['id']),
            ws.cell(row=row, column=2, value=task['name']),
            ws.cell(row=row, column=3, value=task['startDate']),
            ws.cell(row=row, column=4, value=task['endDate']),
            ws.cell(row=row, column=5, value=duration),
            ws.cell(row=row, column=6, value=dependencies_str),
        ]
        
        for cell in cells:
            cell.border = border
            cell.alignment = center_align if cell.column in [1, 5] else left_align
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 30
    
    wb.save(filename)
    print(f"✓ Saved {len(test_tasks)} tasks to {filename}")

if __name__ == '__main__':
    import os
    os.chdir('/var/folders/0j/727fsw6n7b72vdhgshwp50tr0000gn/T/opencode/task-gantt')
    
    print("Generating test data from roadmap...")
    save_as_json('tasks.json')
    save_as_excel('roadmap_test_data.xlsx')
    print("\n✓ Test data generation complete!")
